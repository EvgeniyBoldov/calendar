"""
Сервис импорта плана работ из Excel файла.

Читает Excel файл и извлекает задачи для добавления в план работы.
Настройки парсинга берутся из конфига.
"""
from io import BytesIO
from dataclasses import dataclass
from openpyxl import load_workbook
import xlrd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from ..config import get_settings
from ..models import WorkTask, DataCenter
from ..models.work import TaskStatus


@dataclass
class ImportedTask:
    """Задача, извлечённая из Excel"""
    title: str
    data_center_name: str | None
    estimated_hours: int


@dataclass
class ImportResult:
    """Результат импорта"""
    success: bool
    imported_count: int
    skipped_count: int
    errors: list[str]
    tasks: list[ImportedTask]


class ExcelImportService:
    """Сервис импорта плана работ из Excel"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
        self.settings = get_settings()
    
    def _col_letter_to_index(self, col: str) -> int:
        """Преобразует букву столбца в индекс (A=1, B=2, ...)"""
        col = col.upper()
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def parse_excel(self, file_data: BytesIO) -> ImportResult:
        """
        Парсит Excel файл и извлекает задачи.
        
        Returns:
            ImportResult с извлечёнными задачами
        """
        errors = []
        tasks = []
        
        try:
            wb = load_workbook(file_data, data_only=True)
        except Exception as e_xlsx:
            # Если не удалось открыть как xlsx, пробуем как xls
            try:
                # Для xlrd нужно сбросить указатель файла
                file_data.seek(0)
                file_content = file_data.read()
                workbook = xlrd.open_workbook(file_contents=file_content)
                return self._parse_xls_with_xlrd(workbook)
            except Exception as e_xls:
                return ImportResult(
                    success=False,
                    imported_count=0,
                    skipped_count=0,
                    errors=[f"Не удалось открыть файл: {str(e_xlsx)} / {str(e_xls)}"],
                    tasks=[]
                )
        
        # Ищем нужный лист
        sheet_name = self.settings.excel_import_sheet
        if sheet_name not in wb.sheetnames:
            # Пробуем найти похожий лист
            found = False
            for name in wb.sheetnames:
                if sheet_name.lower() in name.lower():
                    sheet_name = name
                    found = True
                    break
            if not found:
                # Берём первый лист
                sheet_name = wb.sheetnames[0]
                errors.append(f"Лист '{self.settings.excel_import_sheet}' не найден, используется '{sheet_name}'")
        
        ws = wb[sheet_name]
        
        # Индексы столбцов
        desc_col = self._col_letter_to_index(self.settings.excel_import_description_col)
        dc_col = self._col_letter_to_index(self.settings.excel_import_dc_col)
        hours_col = self._col_letter_to_index(self.settings.excel_import_hours_col)
        start_row = self.settings.excel_import_start_row
        
        # Читаем строки
        row_num = start_row
        while True:
            # Читаем описание
            desc_cell = ws.cell(row=row_num, column=desc_col)
            if desc_cell.value is None or str(desc_cell.value).strip() == "":
                # Пустая строка - конец данных
                break
            
            title = str(desc_cell.value).strip()
            
            # Читаем ДЦ
            dc_cell = ws.cell(row=row_num, column=dc_col)
            dc_name = str(dc_cell.value).strip() if dc_cell.value else None
            
            # Читаем часы
            hours_cell = ws.cell(row=row_num, column=hours_col)
            try:
                hours = int(float(hours_cell.value)) if hours_cell.value else 1
                if hours < 1:
                    hours = 1
                if hours > 24:
                    hours = 24
            except (ValueError, TypeError):
                hours = 1
                errors.append(f"Строка {row_num}: не удалось прочитать часы, установлено 1")
            
            tasks.append(ImportedTask(
                title=title,
                data_center_name=dc_name,
                estimated_hours=hours
            ))
            
            row_num += 1
            
            # Защита от бесконечного цикла
            if row_num > 1000:
                errors.append("Достигнут лимит в 1000 строк")
                break
        
        return ImportResult(
            success=True,
            imported_count=len(tasks),
            skipped_count=0,
            errors=errors,
            tasks=tasks
        )
    
    def _parse_xls_with_xlrd(self, wb) -> ImportResult:
        """Парсит .xls файл с помощью xlrd"""
        errors = []
        tasks = []
        
        # Ищем лист
        sheet_name = self.settings.excel_import_sheet
        if sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
        else:
            # Ищем похожий
            found = False
            for name in wb.sheet_names():
                if sheet_name.lower() in name.lower():
                    sheet_name = name
                    found = True
                    break
            
            if found:
                ws = wb.sheet_by_name(sheet_name)
            else:
                ws = wb.sheet_by_index(0)
                errors.append(f"Лист '{self.settings.excel_import_sheet}' не найден, используется '{ws.name}'")

        # Индексы столбцов
        desc_col = self._col_letter_to_index(self.settings.excel_import_description_col) - 1 # xlrd 0-based
        dc_col = self._col_letter_to_index(self.settings.excel_import_dc_col) - 1
        hours_col = self._col_letter_to_index(self.settings.excel_import_hours_col) - 1
        start_row = self.settings.excel_import_start_row - 1 # xlrd 0-based

        # Читаем строки
        for row_idx in range(start_row, ws.nrows):
            # Описание
            try:
                desc_val = ws.cell_value(row_idx, desc_col)
                if not str(desc_val).strip():
                    break
                title = str(desc_val).strip()
            except IndexError:
                break

            # ДЦ
            dc_name = None
            try:
                dc_val = ws.cell_value(row_idx, dc_col)
                if dc_val:
                    dc_name = str(dc_val).strip()
            except IndexError:
                pass

            # Часы
            hours = 1
            try:
                hours_val = ws.cell_value(row_idx, hours_col)
                if hours_val:
                    try:
                        hours = int(float(hours_val))
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_idx + 1}: не удалось прочитать часы, установлено 1")
                if hours < 1: hours = 1
                if hours > 24: hours = 24
            except IndexError:
                pass

            tasks.append(ImportedTask(
                title=title,
                data_center_name=dc_name,
                estimated_hours=hours
            ))
            
            if len(tasks) > 1000:
                errors.append("Достигнут лимит в 1000 строк")
                break

        return ImportResult(
            success=True,
            imported_count=len(tasks),
            skipped_count=0,
            errors=errors,
            tasks=tasks
        )

    async def get_dc_map(self) -> dict[str, str]:
        """Получить маппинг название ДЦ -> ID"""
        result = await self.db.execute(select(DataCenter))
        dcs = result.scalars().all()
        
        dc_map = {}
        for dc in dcs:
            # Добавляем разные варианты названия
            dc_map[dc.name.lower()] = dc.id
            dc_map[dc.name.lower().replace(" ", "")] = dc.id
            # Также добавляем без "ДЦ" префикса
            name_clean = dc.name.lower().replace("дц", "").replace("dc", "").strip()
            if name_clean:
                dc_map[name_clean] = dc.id
        
        return dc_map
    
    async def import_tasks_to_work(
        self,
        work_id: str,
        tasks: list[ImportedTask],
        skip_duplicates: bool = True
    ) -> tuple[int, int, list[str]]:
        """
        Импортирует задачи в работу.
        
        Args:
            work_id: ID работы
            tasks: Список задач для импорта
            skip_duplicates: Пропускать дубликаты (по описанию, ДЦ, часам)
        
        Returns:
            (imported_count, skipped_count, errors)
        """
        errors = []
        imported = 0
        skipped = 0
        
        # Получаем маппинг ДЦ
        dc_map = await self.get_dc_map()
        
        # Получаем существующие задачи работы
        existing_result = await self.db.execute(
            select(WorkTask).where(WorkTask.work_id == work_id)
        )
        existing_tasks = existing_result.scalars().all()
        
        # Создаём set для проверки дубликатов
        existing_set = set()
        for task in existing_tasks:
            key = (
                task.title.lower().strip(),
                task.data_center_id,
                task.estimated_hours
            )
            existing_set.add(key)
        
        # Определяем начальный order
        max_order = max((t.order for t in existing_tasks), default=-1)
        current_order = max_order + 1
        
        for task in tasks:
            # Определяем ID ДЦ
            dc_id = None
            if task.data_center_name:
                dc_name_lower = task.data_center_name.lower().strip()
                dc_id = dc_map.get(dc_name_lower)
                if not dc_id:
                    # Пробуем без пробелов
                    dc_id = dc_map.get(dc_name_lower.replace(" ", ""))
                if not dc_id:
                    errors.append(f"ДЦ '{task.data_center_name}' не найден")
            
            # Проверяем дубликат
            key = (task.title.lower().strip(), dc_id, task.estimated_hours)
            if skip_duplicates and key in existing_set:
                skipped += 1
                continue
            
            # Создаём задачу
            new_task = WorkTask(
                work_id=work_id,
                title=task.title,
                data_center_id=dc_id,
                estimated_hours=task.estimated_hours,
                order=current_order,
                status=TaskStatus.TODO
            )
            self.db.add(new_task)
            existing_set.add(key)
            imported += 1
            current_order += 1
        
        await self.db.flush()
        
        return imported, skipped, errors
